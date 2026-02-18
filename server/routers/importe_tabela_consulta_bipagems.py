"""
Rotas: importe tabela consulta das bipagems em tempo real – importar Excel, deduplicar por "Número de pedido JMS"
(ficando a linha com "Tempo de digitalização" mais recente, ex.: 2026-02-02 14:29:05)
e gravar na coleção pedidos_com_status. GET total e DELETE atuam sobre essa coleção.
"""
import warnings
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO

warnings.filterwarnings("ignore", message="Workbook contains no default style", module="openpyxl")

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from openpyxl import load_workbook
from pymongo.errors import PyMongoError, BulkWriteError

from database import get_db, USER_ID_FIELD
from limiter import limiter
from routers.auth import require_user_id
from table_ids import require_table_id

router = APIRouter(prefix="/importe-tabela-consulta-bipagems", tags=["importe-tabela-consulta-bipagems"])
COLLECTION = "pedidos_com_status"
CHUNK_SIZE = 5000
MAX_CELL_LEN = 50000  # evita documentos enormes e problemas de serialização BSON
IMPORT_DATE_FIELD = "importDate"
HEADER_FLAG = "isHeader"


def _garantir_colecao(db):
    """Cria a coleção no MongoDB se não existir."""
    try:
        if COLLECTION not in db.list_collection_names():
            db.create_collection(COLLECTION)
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao criar coleção no banco de dados: {e}")


def _sanitize_cell(v):
    """Converte valor da célula para string e limita tamanho."""
    if v is None:
        return ""
    s = str(v).strip() if isinstance(v, str) else str(v)
    if len(s) > MAX_CELL_LEN:
        return s[:MAX_CELL_LEN]
    return s


def _parse_tempo_valor(v):
    """Converte valor da coluna 'Tempo de digitalização' para datetime comparável. Falha = datetime mínimo."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return datetime.min.replace(tzinfo=timezone.utc)
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:26], fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return datetime.min.replace(tzinfo=timezone.utc)


def _indices_colunas_obrigatorias(header: list) -> tuple[int, int]:
    """
    Retorna (idx_pedido, idx_tempo) para "Número de pedido JMS" e "Tempo de digitalização".
    Levanta HTTPException 400 se alguma coluna não existir no cabeçalho.
    """
    norm = [str(h).strip().lower() if h is not None else "" for h in header]
    idx_pedido = next((i for i, h in enumerate(norm) if "número de pedido jms" in h or "numero de pedido jms" in h), -1)
    idx_tempo = next((i for i, h in enumerate(norm) if "tempo de digitalização" in h or "tempo de digitalizacao" in h), -1)
    if idx_pedido < 0:
        raise HTTPException(
            status_code=400,
            detail="O ficheiro Excel deve conter a coluna \"Número de pedido JMS\".",
        )
    if idx_tempo < 0:
        raise HTTPException(
            status_code=400,
            detail="O ficheiro Excel deve conter a coluna \"Tempo de digitalização\".",
        )
    return idx_pedido, idx_tempo


# Valor de "Tipo de bipagem" que não deve ser gravado na coleção pedidos_com_status
TIPO_BIPAGEM_EXCLUIR = "assinatura de encomenda"


def _idx_tipo_bipagem(header: list) -> int:
    """Retorna o índice da coluna 'Tipo de bipagem' no cabeçalho, ou -1 se não existir."""
    norm = [str(h).strip().lower() if h is not None else "" for h in header]
    return next((i for i, h in enumerate(norm) if "tipo de bipagem" in h or "tipo bipagem" in h), -1)


def _tipo_bipagem_deve_excluir(valor: str) -> bool:
    """True se o valor de Tipo de bipagem deve excluir a linha do import."""
    v = (valor or "").strip().lower()
    return v == TIPO_BIPAGEM_EXCLUIR.lower()


def _manter_apenas_bipe_mais_recente(rows: list) -> list:
    """
    Agrupa por "Número de pedido JMS" (único por pedido) e mantém só a linha com
    "Tempo de digitalização" mais recente (formato ex.: 2026-02-02 14:29:05).
    Espera rows[0] = cabeçalho; rows[1:] = dados. Retorna [cabeçalho] + linhas deduplicadas.
    """
    if not rows or len(rows) <= 1:
        return rows
    header = rows[0]
    idx_pedido, idx_tempo = _indices_colunas_obrigatorias(header)
    groups = defaultdict(list)
    for row in rows[1:]:
        key = (row[idx_pedido] if idx_pedido < len(row) else "").strip()
        groups[key].append(row)
    chosen = []
    for key, group in groups.items():
        if not group:
            continue
        best = max(group, key=lambda r: _parse_tempo_valor(r[idx_tempo] if idx_tempo < len(r) else None))
        chosen.append(best)
    return [header] + chosen


def _excel_para_linhas(contents: bytes) -> list:
    """Lê a primeira planilha e retorna lista de linhas (cada linha = lista de strings sanitizadas)."""
    wb = load_workbook(filename=BytesIO(contents), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = []
    max_row = ws.max_row if ws.max_row else 0
    max_col = ws.max_column if ws.max_column else 1
    for row_idx in range(1, max_row + 1):
        row = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row.append(_sanitize_cell(cell.value))
        rows.append(row)
    wb.close()
    return rows


def _jms_existentes(col, idx_jms: int, user_id: str) -> set:
    """Retorna o conjunto de 'Número de pedido JMS' já presentes na coleção (docs do usuário com importDate)."""
    if idx_jms < 0:
        return set()
    seen = set()
    q = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
    for doc in col.find(q, {"values": 1}):
        vals = doc.get("values") or []
        if idx_jms < len(vals):
            v = (vals[idx_jms] or "").strip()
            if v:
                seen.add(v)
    return seen


def _insert_batch(col, batch, saved_so_far):
    """Insere um lote; em caso de BulkWriteError devolve mensagem com linha aproximada."""
    try:
        col.insert_many(batch, ordered=True)
    except BulkWriteError as e:
        errs = (e.details or {}).get("writeErrors") or [{}]
        first = errs[0] if errs else {}
        idx = first.get("index", 0)
        approx_row = saved_so_far + idx + 1
        msg = first.get("errmsg", str(e))
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao gravar na linha (aprox.) {approx_row}: {msg}",
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gravar no banco de dados: {e}")


def _parse_datas_query(datas: str | None) -> list[str] | None:
    if not datas or not str(datas).strip():
        return None
    return [d.strip() for d in str(datas).split(",") if d.strip()]


@router.post("")
@limiter.limit("20/minute")
async def importar_pedidos_consultados(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Recebe um arquivo Excel (.xlsx). Modo incremental: não apaga dados anteriores.
    Grava cada linha com importDate (data do envio). Números de pedido JMS já existentes são ignorados.
    Exige colunas "Número de pedido JMS" e "Tempo de digitalização"; mantém uma linha por JMS (mais recente).
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não informado.")
    ext = (file.filename or "").lower()
    if not ext.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    contents = await file.read()
    try:
        rows = _excel_para_linhas(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o Excel: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="O arquivo está vazio ou não tem dados na primeira planilha.")

    rows = _manter_apenas_bipe_mais_recente(rows)
    header = rows[0]
    idx_pedido, _ = _indices_colunas_obrigatorias(header)
    idx_tipo_bipagem = _idx_tipo_bipagem(header)
    data_rows = rows[1:] if len(rows) > 1 else []

    try:
        db = get_db()
        _garantir_colecao(db)
        col = db[COLLECTION]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")

    now = datetime.now(timezone.utc)
    import_date_str = now.strftime("%Y-%m-%d")
    q_user = {USER_ID_FIELD: user_id}

    if col.count_documents(q_user) == 0:
        col.insert_one({**q_user, "values": list(header), HEADER_FLAG: True})

    existing_jms = _jms_existentes(col, idx_pedido, user_id)
    to_insert = []
    for row in data_rows:
        if idx_tipo_bipagem >= 0 and idx_tipo_bipagem < len(row):
            if _tipo_bipagem_deve_excluir(str(row[idx_tipo_bipagem] or "")):
                continue
        jms = (row[idx_pedido] if idx_pedido < len(row) else "").strip()
        if jms and jms in existing_jms:
            continue
        to_insert.append({
            **q_user,
            "values": row,
            "createdAt": now,
            IMPORT_DATE_FIELD: import_date_str,
        })
        if jms:
            existing_jms.add(jms)

    saved = 0
    for i in range(0, len(to_insert), CHUNK_SIZE):
        batch = to_insert[i : i + CHUNK_SIZE]
        _insert_batch(col, batch, saved)
        saved += len(batch)

    return {"saved": saved}


@router.get("/datas")
def listar_datas_importacao(user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """Retorna as datas de importação (importDate) existentes na coleção, ordenadas da mais recente."""
    try:
        db = get_db()
        col = db[COLLECTION]
        q = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
        datas = col.distinct(IMPORT_DATE_FIELD, q)
        datas = sorted([d for d in datas if d], reverse=True)
        return {"datas": datas}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")


@router.get("")
def listar_total_pedidos_consultados(
    datas: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna o total de documentos na coleção (opcionalmente filtrado por datas de importação).
    datas: vírgulas (ex: 2026-02-08,2026-02-09).
    """
    try:
        db = get_db()
        col = db[COLLECTION]
        query = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
        datas_list = _parse_datas_query(datas)
        if datas_list:
            query[IMPORT_DATE_FIELD] = {"$in": datas_list}
        total = col.count_documents(query)
        return {"total": total}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")


@router.delete("")
def excluir_pedidos_consultados(user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """Remove todos os documentos da coleção pedidos_com_status. Requer autenticação."""
    try:
        db = get_db()
        col = db[COLLECTION]
        result = col.delete_many({USER_ID_FIELD: user_id})
        return {"deleted": result.deleted_count}
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao excluir do banco de dados: {e}")
