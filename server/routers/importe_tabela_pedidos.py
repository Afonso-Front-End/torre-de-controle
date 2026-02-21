"""
Rotas: importe tabela de pedidos – recebe arquivo Excel, processa, salva na coleção.
Suporta grandes volumes: leitura em streaming do Excel e inserção em lotes no MongoDB.
Para cada "Número de pedido JMS" é guardada apenas a linha com o "Tempo de digitalização" mais recente.
"""
import warnings
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO

warnings.filterwarnings("ignore", message="Workbook contains no default style", module="openpyxl")

from bson.errors import InvalidId
from bson.objectid import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from openpyxl import load_workbook
from pymongo.errors import PyMongoError, BulkWriteError

from database import get_db, USER_ID_FIELD
from limiter import limiter
from routers.auth import require_user_id
from table_ids import require_table_id
from upload_limits import read_upload_with_limit

router = APIRouter(prefix="/importe-tabela-pedidos", tags=["importe-tabela-pedidos"])
COLLECTION = "pedidos"
CHUNK_SIZE = 5000
MAX_CELL_LEN = 50000  # evita documentos enormes e problemas de serialização BSON
IMPORT_DATE_FIELD = "importDate"  # data do envio (YYYY-MM-DD) para filtrar por data
HEADER_FLAG = "isHeader"  # primeiro doc da coleção = cabeçalho


def _garantir_colecao(db):
    """Cria a coleção no MongoDB se não existir."""
    try:
        if COLLECTION not in db.list_collection_names():
            db.create_collection(COLLECTION)
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao criar coleção no banco de dados: {e}")


def _sanitize_cell(v):
    """Converte valor da célula para string e limita tamanho (evita BSON/doc grande e tipos problemáticos)."""
    if v is None:
        return ""
    s = str(v).strip() if isinstance(v, str) else str(v)
    if len(s) > MAX_CELL_LEN:
        return s[:MAX_CELL_LEN]
    return s


def _parse_tempo_valor(v):
    """Converte valor da coluna 'Tempo de digitalização' para datetime comparável. Falha = datetime mínimo."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return datetime.min.replace(tzinfo=timezone.utc)
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:26], fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return datetime.min.replace(tzinfo=timezone.utc)


def _manter_apenas_bipe_mais_recente(rows: list) -> list:
    """
    Agrupa por "Número de pedido JMS" e mantém só a linha com "Tempo de digitalização" mais recente.
    Espera rows[0] = cabeçalho; rows[1:] = dados. Retorna [cabeçalho] + linhas deduplicadas.
    """
    if not rows or len(rows) <= 1:
        return rows
    header = rows[0]
    norm = [str(h).strip().lower() if h is not None else "" for h in header]
    idx_pedido = next((i for i, h in enumerate(norm) if "número de pedido jms" in h or "numero de pedido jms" in h), -1)
    idx_tempo = next((i for i, h in enumerate(norm) if "tempo de digitalização" in h or "tempo de digitalizacao" in h), -1)
    if idx_pedido < 0 or idx_tempo < 0:
        return rows
    groups = defaultdict(list)
    for row in rows[1:]:
        key = (row[idx_pedido] if idx_pedido < len(row) else "").strip()
        groups[key].append(row)
    chosen = []
    for key, group in groups.items():
        if not group:
            continue
        best = max(group, key=lambda r: _parse_tempo_valor(r[idx_tempo] if idx_tempo < len(r) else None))
        chosen.append(best)
    return [header] + chosen


def _excel_para_linhas(contents: bytes) -> list:
    """Lê a primeira planilha e retorna lista de linhas (cada linha = lista de strings sanitizadas)."""
    wb = load_workbook(filename=BytesIO(contents), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = []
    max_row = ws.max_row if ws.max_row else 0
    max_col = ws.max_column if ws.max_column else 1
    for row_idx in range(1, max_row + 1):
        row = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row.append(_sanitize_cell(cell.value))
        rows.append(row)
    wb.close()
    return rows


def _idx_numero_pedido_jms(header: list) -> int:
    """Retorna o índice da coluna 'Número de pedido JMS' no cabeçalho."""
    norm = [str(h).strip().lower() if h is not None else "" for h in header]
    return next(
        (i for i, h in enumerate(norm) if "número de pedido jms" in h or "numero de pedido jms" in h),
        -1,
    )


def _jms_existentes(col, idx_jms: int, user_id: str) -> set:
    """
    Retorna o conjunto de valores 'Número de pedido JMS' já presentes na coleção
    (apenas documentos de dados do usuário, com importDate). Para deduplicar no import.
    """
    if idx_jms < 0:
        return set()
    seen = set()
    q = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
    for doc in col.find(q, {"values": 1}):
        vals = doc.get("values") or []
        if idx_jms < len(vals):
            v = (vals[idx_jms] or "").strip()
            if v:
                seen.add(v)
    return seen


def _insert_batch(col, batch, saved_so_far):
    """Insere um lote e em caso de erro devolve mensagem com linha aproximada."""
    try:
        col.insert_many(batch, ordered=True)
    except BulkWriteError as e:
        errs = (e.details or {}).get("writeErrors") or [{}]
        first = errs[0] if errs else {}
        idx = first.get("index", 0)
        approx_row = saved_so_far + idx + 1
        msg = first.get("errmsg", str(e))
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao gravar na linha (aprox.) {approx_row}: {msg}",
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gravar no banco de dados: {e}")


@router.post("")
@limiter.limit("20/minute")
async def salvar_pedidos(request: Request, file: UploadFile = File(...), user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """
    Recebe um arquivo Excel (.xlsx). Lê a primeira planilha, sanitiza células e grava em lotes.
    Modo incremental: não apaga dados anteriores. Grava cada linha com importDate (data do envio).
    Números de pedido JMS já existentes no banco são ignorados (não duplicados).
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não informado.")
    ext = (file.filename or "").lower()
    if not ext.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    contents = await read_upload_with_limit(file)
    try:
        rows = _excel_para_linhas(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o Excel: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="O arquivo está vazio ou não tem dados na primeira planilha.")

    rows = _manter_apenas_bipe_mais_recente(rows)
    header = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []

    try:
        db = get_db()
        _garantir_colecao(db)
        col = db[COLLECTION]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")

    now = datetime.now(timezone.utc)
    import_date_str = now.strftime("%Y-%m-%d")
    q_user = {USER_ID_FIELD: user_id}

    # Se a coleção estiver vazia para este usuário, gravar o cabeçalho uma vez (marcado com isHeader).
    if col.count_documents(q_user) == 0:
        col.insert_one({**q_user, "values": list(header), HEADER_FLAG: True})

    idx_jms = _idx_numero_pedido_jms(header)
    existing_jms = _jms_existentes(col, idx_jms, user_id)

    to_insert = []
    for row in data_rows:
        jms = (row[idx_jms] if idx_jms < len(row) else "").strip()
        if jms and jms in existing_jms:
            continue
        to_insert.append({
            **q_user,
            "values": row,
            "createdAt": now,
            IMPORT_DATE_FIELD: import_date_str,
        })
        if jms:
            existing_jms.add(jms)

    saved = 0
    for i in range(0, len(to_insert), CHUNK_SIZE):
        batch = to_insert[i : i + CHUNK_SIZE]
        _insert_batch(col, batch, saved)
        saved += len(batch)

    return {"saved": saved}


def _parse_datas_query(datas: str | None) -> list[str] | None:
    """Converte query param 'datas' (ex: '2026-02-08,2026-02-09') em lista de strings YYYY-MM-DD."""
    if not datas or not str(datas).strip():
        return None
    return [d.strip() for d in str(datas).split(",") if d.strip()]


@router.get("/datas")
def listar_datas_importacao(user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """
    Retorna as datas de importação (importDate) existentes na coleção, ordenadas da mais recente.
    Usado no frontend para o select múltiplo de datas.
    """
    db = get_db()
    col = db[COLLECTION]
    q = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
    datas = col.distinct(IMPORT_DATE_FIELD, q)
    datas = sorted([d for d in datas if d], reverse=True)
    return {"datas": datas}


@router.get("")
def listar_pedidos(
    page: int = 1,
    per_page: int = 100,
    datas: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna pedidos com paginação (skip/limit) para suportar grandes volumes.
    datas: opcional, vírgulas (ex: 2026-02-08,2026-02-09) – exibe apenas registros dessas datas de envio.
    page=1 retorna também o header (primeira linha da coleção).
    """
    per_page = min(max(1, per_page), 500)
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}

    query = dict(q_user)
    datas_list = _parse_datas_query(datas)
    if datas_list:
        query[IMPORT_DATE_FIELD] = {"$in": datas_list}

    # Header: doc com isHeader ou o primeiro doc sem importDate (retrocompatibilidade), do usuário.
    header_doc = col.find_one({**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]}, sort=[("_id", 1)])
    header = list(header_doc.get("values", [])) if header_doc else []
    header_id = header_doc["_id"] if header_doc else None

    # Dados: do usuário; todos exceto o header; opcionalmente filtrar por datas.
    data_query = dict(q_user)
    if header_id:
        data_query["_id"] = {"$ne": header_id}
    if datas_list:
        data_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
    total = col.count_documents(data_query)

    if total == 0:
        return {"data": [], "total": 0, "header": header if page == 1 else None}

    skip = (page - 1) * per_page
    cursor = col.find(data_query).sort("_id", 1).skip(skip).limit(per_page)

    docs = []
    for doc in cursor:
        c = doc.get("createdAt")
        # Garantir serialização correta de datetime
        if c:
            if hasattr(c, "isoformat"):
                created_at = c.isoformat()
            elif isinstance(c, str):
                created_at = c
            else:
                created_at = str(c)
        else:
            created_at = None
        
        docs.append({
            "_id": str(doc["_id"]),
            "values": doc.get("values", []),
            "createdAt": created_at,
            "importDate": doc.get(IMPORT_DATE_FIELD),
        })

    return {"data": docs, "total": total, "header": header if page == 1 else None}


@router.delete("")
def deletar_todos(user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """Remove todos os documentos da coleção pedidos. Requer autenticação."""
    db = get_db()
    col = db[COLLECTION]
    result = col.delete_many({USER_ID_FIELD: user_id})
    return {"deleted": result.deleted_count}


@router.delete("/{doc_id}")
def deletar_linha(doc_id: str, user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """Remove um documento pelo _id. Requer autenticação."""
    try:
        oid = ObjectId(doc_id)
    except InvalidId:
        raise HTTPException(status_code=400, detail="ID inválido.")
    db = get_db()
    col = db[COLLECTION]
    result = col.delete_one({"_id": oid, USER_ID_FIELD: user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro não encontrado.")
    return {"deleted": 1}
