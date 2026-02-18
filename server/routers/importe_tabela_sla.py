"""
Rotas: importe tabela SLA – recebe arquivo Excel, processa, salva na coleção.
Suporta grandes volumes: leitura do Excel e inserção em lotes no MongoDB.
Cálculo de indicadores SLA agrupados por base e por motorista.
"""
import re
import unicodedata
import warnings
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO

warnings.filterwarnings("ignore", message="Workbook contains no default style", module="openpyxl")

from bson.errors import InvalidId
from bson.objectid import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pymongo.errors import BulkWriteError, PyMongoError
from pymongo.operations import InsertOne, UpdateOne

from database import USER_ID_FIELD, get_db
from limiter import limiter
from routers.auth import require_user_id
from table_ids import require_table_id

router = APIRouter(prefix="/importe-tabela-sla", tags=["importe-tabela-sla"])
COLLECTION = "sla_tabela"
COLLECTION_ENTRADA_GALPAO = "entrada_no_galpao"
CHUNK_SIZE = 5000
MAX_CELL_LEN = 50000
IMPORT_DATE_FIELD = "importDate"
HEADER_FLAG = "isHeader"
PERIODO_FIELD = "periodo"  # "AM" | "PM" conforme Horário de saída para entrega
COL_TIPO_BIPAGEM = "tipo de bipagem"
COL_JMS_ENTRADA = "número de pedido jms"  # Coluna para identificar o pedido na entrada no galpão
COL_TEMPO_DIGITALIZACAO = "tempo de digitalização"  # Coluna da entrada no galpão
COL_BASE_ESCANEAMENTO = "base de escaneamento"  # Coluna da entrada no galpão
COL_DIGITALIZADOR = "digitalizador"  # Coluna da entrada no galpão
TIPO_BIPAGEM_EXCLUIR = "entrada no galpão de pacote não expedido"


def _garantir_colecao(db):
    try:
        if COLLECTION not in db.list_collection_names():
            db.create_collection(COLLECTION)
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao criar coleção no banco de dados: {e}")


def _garantir_colecao_entrada_galpao(db):
    try:
        if COLLECTION_ENTRADA_GALPAO not in db.list_collection_names():
            db.create_collection(COLLECTION_ENTRADA_GALPAO)
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao criar coleção no banco de dados: {e}")


def _sanitize_cell(v):
    if v is None:
        return ""
    s = str(v).strip() if isinstance(v, str) else str(v)
    return s[:MAX_CELL_LEN] if len(s) > MAX_CELL_LEN else s


def _excel_para_linhas(contents: bytes) -> list:
    wb = load_workbook(filename=BytesIO(contents), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 1
    for row_idx in range(1, max_row + 1):
        row = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row.append(_sanitize_cell(cell.value))
        rows.append(row)
    wb.close()
    return rows


def _insert_batch(col, batch, saved_so_far):
    try:
        col.insert_many(batch, ordered=True)
    except BulkWriteError as e:
        errs = (e.details or {}).get("writeErrors") or [{}]
        first = errs[0] if errs else {}
        idx = first.get("index", 0)
        approx_row = saved_so_far + idx + 1
        msg = first.get("errmsg", str(e))
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao gravar na linha (aprox.) {approx_row}: {msg}",
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gravar no banco de dados: {e}")


@router.post("")
@limiter.limit("20/minute")
async def salvar_sla(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Recebe um arquivo Excel (.xlsx). Lê a primeira planilha, sanitiza células e grava em lotes.
    Suporta grandes volumes (milhares de linhas). Modo incremental: não apaga dados anteriores.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não informado.")
    ext = (file.filename or "").lower()
    if not ext.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    contents = await file.read()
    try:
        rows = _excel_para_linhas(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o Excel: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="O arquivo está vazio ou não tem dados na primeira planilha.")

    header = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []
    idx_horario = _find_col_index(header, COL_HORARIO_SAIDA)
    idx_jms = _find_col_index(header, COL_JMS)

    try:
        db = get_db()
        _garantir_colecao(db)
        col = db[COLLECTION]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")

    now = datetime.now(timezone.utc)
    import_date_str = now.strftime("%Y-%m-%d")
    q_user = {USER_ID_FIELD: user_id}

    if col.count_documents(q_user) == 0:
        col.insert_one({**q_user, "values": list(header), HEADER_FLAG: True})

    to_insert = []
    for row in data_rows:
        if idx_jms >= 0 and idx_jms < len(row):
            jms_value = str(row[idx_jms] or "").strip()
            if "-" in jms_value:
                continue
        period = None
        if idx_horario >= 0 and idx_horario < len(row):
            period = _parse_time_period(row[idx_horario])
        doc = {
            **q_user,
            "values": row,
            "createdAt": now,
            IMPORT_DATE_FIELD: import_date_str,
        }
        if period:
            doc[PERIODO_FIELD] = period
        to_insert.append(doc)

    saved = 0
    for i in range(0, len(to_insert), CHUNK_SIZE):
        batch = to_insert[i : i + CHUNK_SIZE]
        _insert_batch(col, batch, saved)
        saved += len(batch)

    # Debug: verificar se dados foram salvos corretamente
    import sys
    sys.stderr.write(f"[DEBUG] salvar_sla - user_id: {user_id}, saved: {saved}, total_no_banco: {col.count_documents({USER_ID_FIELD: user_id})}\n")
    sys.stderr.flush()

    return {"saved": saved}


@router.post("/atualizar")
@limiter.limit("20/minute")
async def atualizar_sla(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Recebe um arquivo Excel (.xlsx) igual ao import. Atualiza linhas existentes (por número de pedido JMS)
    e insere as novas (ex.: novos motoristas que receberam pedidos). Não apaga dados.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não informado.")
    ext = (file.filename or "").lower()
    if not ext.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    contents = await file.read()
    try:
        rows = _excel_para_linhas(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o Excel: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="O arquivo está vazio ou não tem dados na primeira planilha.")

    header_row = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []
    idx_horario = _find_col_index(header_row, COL_HORARIO_SAIDA)
    idx_jms = _find_col_index(header_row, COL_JMS)

    try:
        db = get_db()
        _garantir_colecao(db)
        col = db[COLLECTION]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")

    now = datetime.now(timezone.utc)
    import_date_str = now.strftime("%Y-%m-%d")
    q_user = {USER_ID_FIELD: user_id}

    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_doc:
        col.insert_one({**q_user, "values": list(header_row), HEADER_FLAG: True})
        header_doc = col.find_one({**q_user, HEADER_FLAG: True}, sort=[("_id", 1)])
    header_values = list(header_doc.get("values", []))
    idx_jms_db = _find_col_index(header_values, COL_JMS)
    if idx_jms_db < 0:
        idx_jms_db = idx_jms

    # Carregar mapa JMS -> _id numa única agregação (em vez de 1 find_one por linha)
    jms_to_id = {}
    if idx_jms_db >= 0:
        pipeline = [
            {"$match": {USER_ID_FIELD: user_id, "_id": {"$ne": header_doc["_id"]}}},
            {"$project": {"_id": 1, "jms": {"$arrayElemAt": ["$values", idx_jms_db]}}},
        ]
        for doc in col.aggregate(pipeline):
            jms_val = doc.get("jms")
            if jms_val is not None and str(jms_val).strip():
                jms_to_id[str(jms_val).strip()] = doc["_id"]

    # Montar lista de operações (UpdateOne ou InsertOne) para bulk_write
    operations = []
    for row in data_rows:
        jms_value = None
        if idx_jms >= 0 and idx_jms < len(row):
            jms_value = str(row[idx_jms] or "").strip()
            if "-" in jms_value:
                continue
        period = None
        if idx_horario >= 0 and idx_horario < len(row):
            period = _parse_time_period(row[idx_horario])
        doc = {
            **q_user,
            "values": row,
            "createdAt": now,
            IMPORT_DATE_FIELD: import_date_str,
        }
        if period:
            doc[PERIODO_FIELD] = period

        if jms_value is not None and jms_value in jms_to_id:
            set_fields = {"values": row, IMPORT_DATE_FIELD: import_date_str, "updatedAt": now}
            if period is not None:
                set_fields[PERIODO_FIELD] = period
            operations.append(
                UpdateOne(
                    {"_id": jms_to_id[jms_value], USER_ID_FIELD: user_id},
                    {"$set": set_fields},
                )
            )
        else:
            operations.append(InsertOne(doc))

    # Executar em lotes com ordered=False para o servidor poder paralelizar
    for i in range(0, len(operations), CHUNK_SIZE):
        batch = operations[i : i + CHUNK_SIZE]
        try:
            col.bulk_write(batch, ordered=False)
        except BulkWriteError as e:
            errs = ((e.details or {}).get("writeErrors") or [])[:3]
            msg = "; ".join((x.get("errmsg", str(x)) for x in errs)) if errs else str(e)
            raise HTTPException(status_code=500, detail=f"Erro ao gravar em lote: {msg}")
        except PyMongoError as e:
            raise HTTPException(status_code=500, detail=f"Erro ao gravar no banco de dados: {e}")

    updated = sum(1 for op in operations if isinstance(op, UpdateOne))
    inserted = sum(1 for op in operations if isinstance(op, InsertOne))
    return {"updated": updated, "inserted": inserted}


@router.post("/entrada-galpao")
@limiter.limit("20/minute")
async def salvar_entrada_galpao(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Recebe um arquivo Excel (.xlsx) com dados de entrada no galpão.
    Lê a primeira planilha, sanitiza células e grava na coleção entrada_no_galpao.
    Suporta grandes volumes (milhares de linhas). Modo incremental: não apaga dados anteriores.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não informado.")
    ext = (file.filename or "").lower()
    if not ext.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    contents = await file.read()
    try:
        rows = _excel_para_linhas(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o Excel: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="O arquivo está vazio ou não tem dados na primeira planilha.")

    header = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []
    
    # Normalizar header para encontrar colunas
    header_norm = [_normalize_text(h) for h in header]
    col_jms_norm = _normalize_text(COL_JMS_ENTRADA)
    col_tipo_bipagem_norm = _normalize_text(COL_TIPO_BIPAGEM)
    col_tempo_digitalizacao_norm = _normalize_text(COL_TEMPO_DIGITALIZACAO)
    col_base_escaneamento_norm = _normalize_text(COL_BASE_ESCANEAMENTO)
    col_digitalizador_norm = _normalize_text(COL_DIGITALIZADOR)
    
    idx_jms = next((i for i, h in enumerate(header_norm) if col_jms_norm == h), -1)
    idx_tipo_bipagem = next((i for i, h in enumerate(header_norm) if col_tipo_bipagem_norm == h), -1)
    idx_tempo_digitalizacao = next((i for i, h in enumerate(header_norm) if col_tempo_digitalizacao_norm == h), -1)
    idx_base_escaneamento = next((i for i, h in enumerate(header_norm) if col_base_escaneamento_norm == h), -1)
    idx_digitalizador = next((i for i, h in enumerate(header_norm) if col_digitalizador_norm == h), -1)
    
    # Verificar se todas as colunas obrigatórias foram encontradas
    if idx_jms < 0:
        raise HTTPException(status_code=400, detail=f"Coluna '{COL_JMS_ENTRADA}' não encontrada no arquivo.")
    if idx_tipo_bipagem < 0:
        raise HTTPException(status_code=400, detail=f"Coluna '{COL_TIPO_BIPAGEM}' não encontrada no arquivo.")
    if idx_tempo_digitalizacao < 0:
        raise HTTPException(status_code=400, detail=f"Coluna '{COL_TEMPO_DIGITALIZACAO}' não encontrada no arquivo.")
    if idx_base_escaneamento < 0:
        raise HTTPException(status_code=400, detail=f"Coluna '{COL_BASE_ESCANEAMENTO}' não encontrada no arquivo.")
    if idx_digitalizador < 0:
        raise HTTPException(status_code=400, detail=f"Coluna '{COL_DIGITALIZADOR}' não encontrada no arquivo.")

    try:
        db = get_db()
        _garantir_colecao_entrada_galpao(db)
        col = db[COLLECTION_ENTRADA_GALPAO]
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao banco de dados: {e}")

    now = datetime.now(timezone.utc)
    import_date_str = now.strftime("%Y-%m-%d")
    q_user = {USER_ID_FIELD: user_id}

    # Criar novo header apenas com as colunas que queremos salvar
    novo_header = [
        header[idx_jms],
        header[idx_tipo_bipagem],
        header[idx_tempo_digitalizacao],
        header[idx_base_escaneamento],
        header[idx_digitalizador],
    ]
    
    # Índices no novo array (ordem: JMS, Tipo bipagem, Tempo digitalização, Base escaneamento, Digitalizador)
    novo_idx_jms = 0
    novo_idx_tipo_bipagem = 1
    novo_idx_tempo_digitalizacao = 2
    novo_idx_base_escaneamento = 3
    novo_idx_digitalizador = 4

    if col.count_documents(q_user) == 0:
        col.insert_one({**q_user, "values": novo_header, HEADER_FLAG: True})

    to_insert = []
    for row in data_rows:
        # Criar novo array apenas com os valores das colunas que queremos salvar
        novo_row = [
            _get(row, idx_jms),
            _get(row, idx_tipo_bipagem),
            _get(row, idx_tempo_digitalizacao),
            _get(row, idx_base_escaneamento),
            _get(row, idx_digitalizador),
        ]
        doc = {
            **q_user,
            "values": novo_row,
            "createdAt": now,
            IMPORT_DATE_FIELD: import_date_str,
        }
        to_insert.append(doc)

    saved = 0
    for i in range(0, len(to_insert), CHUNK_SIZE):
        batch = to_insert[i : i + CHUNK_SIZE]
        _insert_batch(col, batch, saved)
        saved += len(batch)

    return {"saved": saved}


def _parse_csv_param(param: str | None) -> list[str] | None:
    """Converte query param em vírgulas (ex: datas=2026-01-01,2026-01-02) em lista."""
    if not param or not str(param).strip():
        return None
    return [p.strip() for p in str(param).split(",") if p.strip()]


# Colunas usadas no cálculo de SLA (nome normalizado para match no header)
COL_BASE = "base de entrega"
COL_MOTORISTA = "responsável pela entrega"
COL_MARCA = "marca de assinatura"
COL_HORARIO_SAIDA = "horário de saída para entrega"
COL_CIDADE_DESTINO = "cidade destino"
COL_JMS = "número de pedido jms"

MARCAS_ENTREGUE = ("recebimento com assinatura normal", "assinatura de devolução")
MARCA_NAO_ENTREGUE = "nao entregue"


def _normalize_text(s: str) -> str:
    """Lowercase e remove acentos para comparação de headers e da coluna Marca."""
    if not s:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_col_index(header: list, col_name: str) -> int:
    """Retorna o índice da coluna cujo header contém col_name (ex: 'marca de assinatura')."""
    col_norm = _normalize_text(col_name)
    for i, h in enumerate(header):
        if col_norm in _normalize_text(h):
            return i
    return -1


def _get(values: list, idx: int) -> str:
    if idx < 0 or idx >= len(values):
        return ""
    v = values[idx]
    return str(v).strip() if v is not None else ""


def _normalize_marca_val(cell_value: str) -> str:
    """Normaliza o valor da célula 'Marca de assinatura' para comparação."""
    return _normalize_text(cell_value or "")


def _parse_time_period(cell_value) -> str | None:
    """
    Extrai período AM/PM do valor da célula "Horário de saída para entrega".
    AM = 00:00–11:59, PM = 12:00–23:59. Retorna None se não conseguir parsear.
    """
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return None
    hour = None
    if isinstance(cell_value, (int, float)):
        if 0 <= cell_value < 1:  # Excel serial time
            hour = int((cell_value * 24) % 24)
    elif isinstance(cell_value, datetime):
        hour = cell_value.hour
    elif isinstance(cell_value, str):
        s = str(cell_value).strip()
        # Float como string
        try:
            v = float(s)
            if 0 <= v < 1:
                hour = int((v * 24) % 24)
        except ValueError:
            pass
        if hour is None:
            # "HH:MM" ou "HH:MM:SS" ou "HH:MM:SS.xxx"
            m = re.match(r"^(\d{1,2})\s*:\s*(\d{2})(?:\s*:\s*\d{2})?(?:\.\d+)?", s)
            if m:
                hour = int(m.group(1)) % 24
            else:
                for fmt in ("%H:%M", "%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M"):
                    try:
                        dt = datetime.strptime(s[:19], fmt)
                        hour = dt.hour
                        break
                    except ValueError:
                        continue
    if hour is None:
        return None
    return "AM" if 0 <= hour < 12 else "PM"


def _parse_time_to_minutes(cell_value) -> int | None:
    """
    Converte um valor de horário (célula Excel ou string) para minutos desde meia-noite.
    Retorna None se não conseguir parsear.
    """
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return None
    hour = None
    minute = 0
    if isinstance(cell_value, (int, float)):
        if 0 <= cell_value < 1:  # Excel serial time
            total_minutes = int(cell_value * 24 * 60)
            hour = total_minutes // 60
            minute = total_minutes % 60
        else:
            return None
    elif isinstance(cell_value, datetime):
        hour = cell_value.hour
        minute = cell_value.minute
    elif isinstance(cell_value, str):
        s = str(cell_value).strip()
        # Float como string
        try:
            v = float(s)
            if 0 <= v < 1:
                total_minutes = int(v * 24 * 60)
                hour = total_minutes // 60
                minute = total_minutes % 60
        except ValueError:
            pass
        if hour is None:
            # "HH:MM" ou "HH:MM:SS" ou "HH:MM:SS.xxx"
            m = re.match(r"^(\d{1,2})\s*:\s*(\d{2})(?:\s*:\s*(\d{2}))?(?:\.\d+)?", s)
            if m:
                hour = int(m.group(1)) % 24
                minute = int(m.group(2)) % 60
            else:
                for fmt in ("%H:%M", "%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M"):
                    try:
                        dt = datetime.strptime(s[:19], fmt)
                        hour = dt.hour
                        minute = dt.minute
                        break
                    except ValueError:
                        continue
    if hour is None:
        return None
    return hour * 60 + minute


def _compare_times(time1_value, time2_value) -> int | None:
    """
    Compara dois valores de horário.
    Retorna:
    - -1 se time1 < time2
    - 0 se time1 == time2
    - 1 se time1 > time2
    - None se algum não puder ser parseado
    """
    minutes1 = _parse_time_to_minutes(time1_value)
    minutes2 = _parse_time_to_minutes(time2_value)
    if minutes1 is None or minutes2 is None:
        return None
    if minutes1 < minutes2:
        return -1
    elif minutes1 > minutes2:
        return 1
    else:
        return 0


@router.get("/indicadores")
def indicadores_sla(
    datas: str | None = None,
    bases: str | None = None,
    cidades: str | None = None,
    periodo: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Calcula indicadores de SLA agrupados por Base e por Motorista (Responsável pela entrega).
    Considera "entregue" as linhas com Marca de assinatura = Recebimento com assinatura normal ou Assinatura de devolução.
    bases: opcional, vírgulas – filtra apenas linhas cuja Base de entrega está na lista (afeta porMotorista).
    cidades: opcional, vírgulas – filtra apenas linhas cuja Cidade Destino está na lista (quando existe coluna).
    periodo: opcional – "AM" ou "PM" para filtrar por período (Horário de saída para entrega); omitir = Todos.
    """
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}
    datas_list = _parse_csv_param(datas)
    bases_list = _parse_csv_param(bases)
    cidades_list_raw = _parse_csv_param(cidades)
    
    # Normalizar cidades para comparação (case-insensitive, sem acentos)
    cidades_list = None
    if cidades_list_raw:
        cidades_list = [_normalize_text(c) for c in cidades_list_raw]

    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_doc:
        return {"header": [], "porBase": [], "porMotorista": []}

    header = list(header_doc.get("values", []))
    header_id = header_doc["_id"]

    idx_base = _find_col_index(header, COL_BASE)
    idx_motorista = _find_col_index(header, COL_MOTORISTA)
    idx_marca = _find_col_index(header, COL_MARCA)
    idx_cidade = _find_col_index(header, COL_CIDADE_DESTINO)
    idx_jms = _find_col_index(header, COL_JMS)
    idx_horario_saida = _find_col_index(header, COL_HORARIO_SAIDA)

    if idx_marca < 0 or idx_base < 0 or idx_motorista < 0:
        return {"header": [], "porBase": [], "porMotorista": []}

    data_query = {**q_user, "_id": {"$ne": header_id}}
    if datas_list:
        data_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
    if periodo and periodo.strip().upper() in ("AM", "PM"):
        data_query[PERIODO_FIELD] = periodo.strip().upper()

    por_base = defaultdict(lambda: {"totalEntregues": 0, "naoEntregues": 0})
    # Agregação por (motorista, base): somamos stats e reunimos o set de cidades (Cidade Destino)
    por_motorista_base = defaultdict(lambda: {"totalEntregues": 0, "naoEntregues": 0, "cidades": set(), "entradasGalpao": 0})
    
    # Buscar todos os pedidos da coleção entrada_no_galpao que devem ser excluídos
    # IMPORTANTE: Usar sempre as mesmas datas do filtro da SLA
    # Mapa: número_pedido_jms -> tempo_de_digitalizacao (para comparar com horário de saída)
    pedidos_excluir = {}
    try:
        col_entrada = db[COLLECTION_ENTRADA_GALPAO]
        header_entrada_doc = col_entrada.find_one(
            {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
            sort=[("_id", 1)],
        )
        if header_entrada_doc:
            header_entrada = list(header_entrada_doc.get("values", []))
            header_entrada_id = header_entrada_doc["_id"]
            header_entrada_norm = [_normalize_text(h) for h in header_entrada]
            col_jms_entrada_norm = _normalize_text(COL_JMS_ENTRADA)
            col_tipo_bipagem_entrada_norm = _normalize_text(COL_TIPO_BIPAGEM)
            col_tempo_digitalizacao_norm = _normalize_text(COL_TEMPO_DIGITALIZACAO)
            idx_jms_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_jms_entrada_norm == h), -1)
            idx_tipo_bipagem_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tipo_bipagem_entrada_norm == h), -1)
            idx_tempo_digitalizacao_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tempo_digitalizacao_norm == h), -1)
            
            if idx_jms_entrada >= 0 and idx_tipo_bipagem_entrada >= 0:
                # Sempre usar as mesmas datas do filtro da SLA
                entrada_query = {**q_user, "_id": {"$ne": header_entrada_id}}
                if datas_list:
                    entrada_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
                # Se não há filtro de datas na SLA, não filtrar entrada no galpão também (buscar todas)
                # Ordenar por _id desc para pegar o mais recente primeiro
                cursor_entrada = col_entrada.find(entrada_query, {"values": 1}).sort("_id", -1)
                # Agrupar por JMS para pegar apenas o mais recente
                candidatos_por_jms = {}
                for doc_entrada in cursor_entrada:
                    vals_entrada = doc_entrada.get("values") or []
                    jms_entrada = _get(vals_entrada, idx_jms_entrada) if idx_jms_entrada >= 0 else ""
                    tipo_bipagem = _get(vals_entrada, idx_tipo_bipagem_entrada) if idx_tipo_bipagem_entrada >= 0 else ""
                    if jms_entrada and tipo_bipagem:
                        jms_entrada_norm = str(jms_entrada).strip()
                        tipo_bipagem_norm = _normalize_text(str(tipo_bipagem))
                        if tipo_bipagem_norm == _normalize_text(TIPO_BIPAGEM_EXCLUIR):
                            # Agrupar por JMS (já ordenado por _id desc, então o primeiro será o mais recente)
                            if jms_entrada_norm not in candidatos_por_jms:
                                tempo_digitalizacao = _get(vals_entrada, idx_tempo_digitalizacao_entrada) if idx_tempo_digitalizacao_entrada >= 0 else None
                                # Armazenar o tempo de digitalização do mais recente para comparação posterior
                                pedidos_excluir[jms_entrada_norm] = tempo_digitalizacao
                                candidatos_por_jms[jms_entrada_norm] = True
    except Exception:
        # Se houver erro ao acessar a coleção de entrada, continuar sem excluir pedidos
        pass

    # Se há filtro de cidades, primeiro coletar todas as cidades únicas disponíveis nos dados
    # para verificar se todas estão selecionadas (otimização: se todas selecionadas, não filtrar)
    if cidades_list and idx_cidade >= 0:
        todas_cidades_disponiveis = set()
        todas_cidades_originais = {}  # Mapa cidade_norm -> cidade_original para logs
        temp_cursor = col.find(data_query, {"values": 1})
        for doc in temp_cursor:
            vals = doc.get("values") or []
            base = _get(vals, idx_base) or "(sem base)"
            if bases_list and base not in bases_list:
                continue
            cidade = _get(vals, idx_cidade) if idx_cidade >= 0 else ""
            cidade_original = (cidade or "").strip()
            cidade_norm = _normalize_text(cidade_original) if cidade_original else ""
            if cidade_norm:  # Só adicionar se não estiver vazia
                todas_cidades_disponiveis.add(cidade_norm)
                todas_cidades_originais[cidade_norm] = cidade_original
        # Se todas as cidades disponíveis estão na lista de filtro, tratar como sem filtro
        cidades_set = set(cidades_list)
        todas_cidades_disponiveis_set = todas_cidades_disponiveis
        
        # Verificar se TODAS as cidades disponíveis estão na lista recebida
        # E se a lista recebida não tem cidades extras (ou seja, são exatamente as mesmas)
        todas_selecionadas = (
            todas_cidades_disponiveis_set 
            and len(todas_cidades_disponiveis_set) > 0
            and todas_cidades_disponiveis_set.issubset(cidades_set)
            and len(cidades_set) == len(todas_cidades_disponiveis_set)  # Garantir que não há extras
        )
        
        if todas_selecionadas:
            cidades_list = None  # Otimização: não filtrar se todas estão selecionadas

    cursor = col.find(data_query, {"values": 1})
    contador_total = 0
    contador_filtrado = 0
    for doc in cursor:
        vals = doc.get("values") or []
        marca = _normalize_marca_val(_get(vals, idx_marca))
        base = _get(vals, idx_base) or "(sem base)"
        if bases_list and base not in bases_list:
            continue
        motorista = _get(vals, idx_motorista) or "(sem motorista)"
        cidade = _get(vals, idx_cidade) if idx_cidade >= 0 else ""
        cidade_original = (cidade or "").strip()
        cidade_norm = _normalize_text(cidade_original) if cidade_original else ""  # Normalizar para comparação
        contador_total += 1
        # Filtrar por cidades apenas se houver filtro e a cidade não estiver na lista
        if cidades_list and cidade_norm not in cidades_list:
            continue
        contador_filtrado += 1
        
        key_mb = (motorista, base)
        
        # Verificar se o pedido deve ser excluído (está na entrada no galpão com tipo específico)
        # Mas só excluir se o horário de saída não existir ou for antes/igual ao tempo de digitalização
        jms_value = ""
        deve_excluir = False
        if idx_jms >= 0:
            jms_value = str(_get(vals, idx_jms) or "").strip()
            if jms_value and jms_value in pedidos_excluir:
                tempo_digitalizacao = pedidos_excluir[jms_value]
                horario_saida = _get(vals, idx_horario_saida) if idx_horario_saida >= 0 else None
                
                # Se não há horário de saída, excluir (pedido não saiu para entrega)
                # Se há horário de saída, comparar com tempo de digitalização
                if horario_saida is None or not str(horario_saida).strip():
                    deve_excluir = True
                else:
                    # Comparar horários: se horário de saída é depois do tempo de digitalização, NÃO excluir
                    comparacao = _compare_times(horario_saida, tempo_digitalizacao)
                    if comparacao is None:
                        # Se não conseguiu comparar, excluir por segurança
                        deve_excluir = True
                    elif comparacao <= 0:
                        # Horário de saída é antes ou igual ao tempo de digitalização, excluir
                        deve_excluir = True
                    # Se comparacao > 0 (horário de saída é depois), não excluir
                
                if deve_excluir:
                    por_motorista_base[key_mb]["entradasGalpao"] += 1
        
        # Se deve excluir, não contar no SLA
        if deve_excluir:
            continue
        # Adicionar cidade original (não normalizada) ao set para exibição (apenas se não estiver vazia)
        if cidade_original:
            por_motorista_base[key_mb]["cidades"].add(cidade_original)
        if marca == MARCA_NAO_ENTREGUE:
            por_base[base]["naoEntregues"] += 1
            por_motorista_base[key_mb]["naoEntregues"] += 1
        elif marca in MARCAS_ENTREGUE:
            por_base[base]["totalEntregues"] += 1
            por_motorista_base[key_mb]["totalEntregues"] += 1
    

    def _pct_sla(total_entregues: int, nao_entregues: int) -> float:
        total = total_entregues + nao_entregues
        return round((total_entregues / total * 100), 1) if total > 0 else 0.0

    def build_list(d):
        return [
            {
                "nome": nome,
                "total": stats["totalEntregues"] + stats["naoEntregues"],
                "totalEntregues": stats["totalEntregues"],
                "naoEntregues": stats["naoEntregues"],
                "percentualSla": _pct_sla(stats["totalEntregues"], stats["naoEntregues"]),
            }
            for nome, stats in sorted(d.items())
        ]

    def build_list_motorista_base():
        out = []
        for (motorista, base), data in sorted(por_motorista_base.items()):
            cidades_list_sorted = sorted(data["cidades"]) if data["cidades"] else []
            out.append({
                "nome": motorista,
                "base": base,
                "total": data["totalEntregues"] + data["naoEntregues"],
                "totalEntregues": data["totalEntregues"],
                "naoEntregues": data["naoEntregues"],
                "percentualSla": _pct_sla(data["totalEntregues"], data["naoEntregues"]),
                "cidades": cidades_list_sorted,
                "entradasGalpao": data.get("entradasGalpao", 0),
            })
        return out

    resp_header = ["Motorista", "Base de entrega", "Total entregues", "Não entregues", "Total", "% SLA", "Entrada do galpao"]

    return {
        "header": resp_header,
        "porBase": build_list(por_base),
        "porMotorista": build_list_motorista_base(),
    }


@router.get("/nao-entregues")
def listar_nao_entregues_motorista(
    motorista: str,
    base: str,
    datas: str | None = None,
    cidades: str | None = None,
    periodo: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna todos os pedidos não entregues de um motorista numa base.
    motorista e base são obrigatórios. Filtros opcionais: datas, cidades, periodo (AM/PM).
    """
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}
    datas_list = _parse_csv_param(datas)
    cidades_list_raw = _parse_csv_param(cidades)
    # Normalizar cidades para comparação (case-insensitive, sem acentos)
    cidades_list = None
    if cidades_list_raw:
        cidades_list = [_normalize_text(c) for c in cidades_list_raw]

    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_doc:
        return {"data": [], "header": []}

    header = list(header_doc.get("values", []))
    header_id = header_doc["_id"]
    idx_base = _find_col_index(header, COL_BASE)
    idx_motorista = _find_col_index(header, COL_MOTORISTA)
    idx_marca = _find_col_index(header, COL_MARCA)
    idx_cidade = _find_col_index(header, COL_CIDADE_DESTINO)
    idx_jms = _find_col_index(header, COL_JMS)
    idx_horario_saida = _find_col_index(header, COL_HORARIO_SAIDA)

    if idx_marca < 0 or idx_base < 0 or idx_motorista < 0:
        return {"data": [], "header": header}

    data_query = {**q_user, "_id": {"$ne": header_id}}
    if periodo and periodo.strip().upper() in ("AM", "PM"):
        data_query[PERIODO_FIELD] = periodo.strip().upper()
    if datas_list:
        data_query[IMPORT_DATE_FIELD] = {"$in": datas_list}

    motorista_norm = (motorista or "").strip() or "(sem motorista)"
    base_norm = (base or "").strip() or "(sem base)"
    
    # Buscar pedidos da entrada no galpão que devem ser excluídos (mesma lógica do indicadores_sla)
    pedidos_excluir = {}
    try:
        col_entrada = db[COLLECTION_ENTRADA_GALPAO]
        header_entrada_doc = col_entrada.find_one(
            {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
            sort=[("_id", 1)],
        )
        if header_entrada_doc:
            header_entrada = list(header_entrada_doc.get("values", []))
            header_entrada_id = header_entrada_doc["_id"]
            header_entrada_norm = [_normalize_text(h) for h in header_entrada]
            col_jms_entrada_norm = _normalize_text(COL_JMS_ENTRADA)
            col_tipo_bipagem_entrada_norm = _normalize_text(COL_TIPO_BIPAGEM)
            col_tempo_digitalizacao_norm = _normalize_text(COL_TEMPO_DIGITALIZACAO)
            idx_jms_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_jms_entrada_norm == h), -1)
            idx_tipo_bipagem_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tipo_bipagem_entrada_norm == h), -1)
            idx_tempo_digitalizacao_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tempo_digitalizacao_norm == h), -1)
            
            if idx_jms_entrada >= 0 and idx_tipo_bipagem_entrada >= 0:
                entrada_query = {**q_user, "_id": {"$ne": header_entrada_id}}
                if datas_list:
                    entrada_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
                # Ordenar por _id desc para pegar o mais recente primeiro
                cursor_entrada = col_entrada.find(entrada_query, {"values": 1}).sort("_id", -1)
                # Agrupar por JMS para pegar apenas o mais recente
                candidatos_por_jms = {}
                for doc_entrada in cursor_entrada:
                    vals_entrada = doc_entrada.get("values") or []
                    jms_entrada = _get(vals_entrada, idx_jms_entrada) if idx_jms_entrada >= 0 else ""
                    tipo_bipagem = _get(vals_entrada, idx_tipo_bipagem_entrada) if idx_tipo_bipagem_entrada >= 0 else ""
                    if jms_entrada and tipo_bipagem:
                        jms_entrada_norm = str(jms_entrada).strip()
                        tipo_bipagem_norm = _normalize_text(str(tipo_bipagem))
                        if tipo_bipagem_norm == _normalize_text(TIPO_BIPAGEM_EXCLUIR):
                            # Agrupar por JMS (já ordenado por _id desc, então o primeiro será o mais recente)
                            if jms_entrada_norm not in candidatos_por_jms:
                                tempo_digitalizacao = _get(vals_entrada, idx_tempo_digitalizacao_entrada) if idx_tempo_digitalizacao_entrada >= 0 else None
                                pedidos_excluir[jms_entrada_norm] = tempo_digitalizacao
                                candidatos_por_jms[jms_entrada_norm] = True
    except Exception:
        pass

    docs = []
    cursor = col.find(data_query, {"values": 1, IMPORT_DATE_FIELD: 1, "createdAt": 1})
    for doc in cursor:
        vals = doc.get("values") or []
        if _normalize_text(_get(vals, idx_motorista)) != _normalize_text(motorista_norm):
            continue
        if _normalize_text(_get(vals, idx_base)) != _normalize_text(base_norm):
            continue
        if _normalize_marca_val(_get(vals, idx_marca)) != MARCA_NAO_ENTREGUE:
            continue
        if cidades_list and idx_cidade >= 0:
            cidade_norm = _normalize_text((_get(vals, idx_cidade) or "").strip()) if (_get(vals, idx_cidade) or "").strip() else ""
            if cidade_norm not in cidades_list:
                continue
        # Excluir pedidos que estão na entrada no galpão
        if idx_jms >= 0:
            jms_value = str(_get(vals, idx_jms) or "").strip()
            if jms_value and jms_value in pedidos_excluir:
                continue
        c = doc.get("createdAt")
        docs.append({
            "_id": str(doc["_id"]),
            "values": vals,
            "createdAt": c.isoformat() if c else None,
            "importDate": doc.get(IMPORT_DATE_FIELD),
        })

    return {"data": docs, "header": header}


@router.get("/entrada-galpao")
def listar_entrada_galpao_motorista(
    motorista: str,
    base: str,
    datas: str | None = None,
    cidades: str | None = None,
    periodo: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna todos os pedidos de entrada no galpão (não expedido) de um motorista numa base.
    Retorna dados diretamente da coleção entrada_no_galpao com suas próprias colunas.
    motorista e base são obrigatórios. Filtros opcionais: datas, cidades, periodo (AM/PM).
    """
    db = get_db()
    col_sla = db[COLLECTION]
    col_entrada = db[COLLECTION_ENTRADA_GALPAO]
    q_user = {USER_ID_FIELD: user_id}
    datas_list = _parse_csv_param(datas)
    cidades_list_raw = _parse_csv_param(cidades)
    cidades_list = None
    if cidades_list_raw:
        cidades_list = [_normalize_text(c) for c in cidades_list_raw]

    # Buscar header da coleção entrada_no_galpao
    header_entrada_doc = col_entrada.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_entrada_doc:
        return {"data": [], "header": []}

    header_entrada = list(header_entrada_doc.get("values", []))
    header_entrada_id = header_entrada_doc["_id"]
    header_entrada_norm = [_normalize_text(h) for h in header_entrada]
    
    # Encontrar índices das colunas na entrada_no_galpao
    col_jms_entrada_norm = _normalize_text(COL_JMS_ENTRADA)
    col_tipo_bipagem_entrada_norm = _normalize_text(COL_TIPO_BIPAGEM)
    col_tempo_digitalizacao_norm = _normalize_text(COL_TEMPO_DIGITALIZACAO)
    idx_jms_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_jms_entrada_norm == h), -1)
    idx_tipo_bipagem_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tipo_bipagem_entrada_norm == h), -1)
    idx_tempo_digitalizacao_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tempo_digitalizacao_norm == h), -1)
    
    if idx_jms_entrada < 0 or idx_tipo_bipagem_entrada < 0:
        return {"data": [], "header": header_entrada}

    # Buscar header da SLA para encontrar motorista e base (para filtrar)
    header_sla_doc = col_sla.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_sla_doc:
        return {"data": [], "header": header_entrada}
    
    header_sla = list(header_sla_doc.get("values", []))
    idx_base_sla = _find_col_index(header_sla, COL_BASE)
    idx_motorista_sla = _find_col_index(header_sla, COL_MOTORISTA)
    idx_cidade_sla = _find_col_index(header_sla, COL_CIDADE_DESTINO)
    idx_jms_sla = _find_col_index(header_sla, COL_JMS)
    idx_horario_saida_sla = _find_col_index(header_sla, COL_HORARIO_SAIDA)
    
    if idx_base_sla < 0 or idx_motorista_sla < 0 or idx_jms_sla < 0:
        return {"data": [], "header": header_entrada}

    motorista_norm = (motorista or "").strip() or "(sem motorista)"
    base_norm = (base or "").strip() or "(sem base)"

    # Buscar todos os JMS do motorista/base na SLA com seus horários de saída (para filtrar entrada no galpão)
    jms_motorista_base = {}  # Mapa: jms -> horario_saida
    data_query_sla = {**q_user, "_id": {"$ne": header_sla_doc["_id"]}}
    if periodo and periodo.strip().upper() in ("AM", "PM"):
        data_query_sla[PERIODO_FIELD] = periodo.strip().upper()
    if datas_list:
        data_query_sla[IMPORT_DATE_FIELD] = {"$in": datas_list}
    
    cursor_sla = col_sla.find(data_query_sla, {"values": 1})
    for doc_sla in cursor_sla:
        vals_sla = doc_sla.get("values") or []
        if _normalize_text(_get(vals_sla, idx_motorista_sla)) != _normalize_text(motorista_norm):
            continue
        if _normalize_text(_get(vals_sla, idx_base_sla)) != _normalize_text(base_norm):
            continue
        if cidades_list and idx_cidade_sla >= 0:
            cidade_norm = _normalize_text((_get(vals_sla, idx_cidade_sla) or "").strip()) if (_get(vals_sla, idx_cidade_sla) or "").strip() else ""
            if cidade_norm not in cidades_list:
                continue
        jms_sla = str(_get(vals_sla, idx_jms_sla) or "").strip()
        if jms_sla:
            horario_saida = _get(vals_sla, idx_horario_saida_sla) if idx_horario_saida_sla >= 0 else None
            jms_motorista_base[jms_sla] = horario_saida

    # Buscar pedidos da entrada_no_galpao que correspondem ao motorista/base
    entrada_query = {**q_user, "_id": {"$ne": header_entrada_id}}
    if datas_list:
        entrada_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
    
    # Coletar todos os documentos da entrada_no_galpao que atendem os critérios básicos
    candidatos_por_jms = {}  # Mapa: jms -> lista de documentos candidatos
    cursor_entrada = col_entrada.find(entrada_query, {"values": 1, IMPORT_DATE_FIELD: 1, "createdAt": 1}).sort("_id", -1)
    for doc_entrada in cursor_entrada:
        vals_entrada = doc_entrada.get("values") or []
        jms_entrada = str(_get(vals_entrada, idx_jms_entrada) or "").strip() if idx_jms_entrada >= 0 else ""
        tipo_bipagem = _get(vals_entrada, idx_tipo_bipagem_entrada) if idx_tipo_bipagem_entrada >= 0 else ""
        
        # Filtrar apenas os que têm tipo correto e JMS do motorista/base
        if not jms_entrada or not tipo_bipagem:
            continue
        tipo_bipagem_norm = _normalize_text(str(tipo_bipagem))
        if tipo_bipagem_norm != _normalize_text(TIPO_BIPAGEM_EXCLUIR):
            continue
        if jms_entrada not in jms_motorista_base:
            continue
        
        # Agrupar por JMS (já ordenado por _id desc, então o primeiro será o mais recente)
        if jms_entrada not in candidatos_por_jms:
            candidatos_por_jms[jms_entrada] = []
        candidatos_por_jms[jms_entrada].append(doc_entrada)
    
    # Para cada JMS, pegar apenas o mais recente e validar horário
    docs = []
    for jms_entrada, candidatos in candidatos_por_jms.items():
        if not candidatos:
            continue
        
        # Pegar o mais recente (primeiro da lista, já que está ordenado por _id desc)
        doc_entrada = candidatos[0]
        vals_entrada = doc_entrada.get("values") or []
        tempo_digitalizacao = _get(vals_entrada, idx_tempo_digitalizacao_entrada) if idx_tempo_digitalizacao_entrada >= 0 else None
        
        # Validar horário: se o horário de saída for depois do tempo de digitalização, não incluir
        horario_saida = jms_motorista_base[jms_entrada]
        if horario_saida is not None and str(horario_saida).strip():
            # Comparar horários: se horário de saída é depois do tempo de digitalização, não incluir
            comparacao = _compare_times(horario_saida, tempo_digitalizacao)
            if comparacao is not None and comparacao > 0:
                # Horário de saída é depois do tempo de digitalização, pedido saiu para entrega
                continue
        
        c = doc_entrada.get("createdAt")
        docs.append({
            "_id": str(doc_entrada["_id"]),
            "values": vals_entrada,
            "createdAt": c.isoformat() if c else None,
            "importDate": doc_entrada.get(IMPORT_DATE_FIELD),
        })

    return {"data": docs, "header": header_entrada}


@router.get("/entregues")
def listar_entregues_motorista(
    motorista: str,
    base: str,
    datas: str | None = None,
    cidades: str | None = None,
    periodo: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna todos os pedidos entregues de um motorista numa base.
    motorista e base são obrigatórios. Filtros opcionais: datas, cidades, periodo (AM/PM).
    """
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}
    datas_list = _parse_csv_param(datas)
    cidades_list_raw = _parse_csv_param(cidades)
    # Normalizar cidades para comparação (case-insensitive, sem acentos)
    cidades_list = None
    if cidades_list_raw:
        cidades_list = [_normalize_text(c) for c in cidades_list_raw]

    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_doc:
        return {"data": [], "header": []}

    header = list(header_doc.get("values", []))
    header_id = header_doc["_id"]
    idx_base = _find_col_index(header, COL_BASE)
    idx_motorista = _find_col_index(header, COL_MOTORISTA)
    idx_marca = _find_col_index(header, COL_MARCA)
    idx_cidade = _find_col_index(header, COL_CIDADE_DESTINO)
    idx_jms = _find_col_index(header, COL_JMS)
    idx_horario_saida = _find_col_index(header, COL_HORARIO_SAIDA)

    if idx_marca < 0 or idx_base < 0 or idx_motorista < 0:
        return {"data": [], "header": header}

    data_query = {**q_user, "_id": {"$ne": header_id}}
    if periodo and periodo.strip().upper() in ("AM", "PM"):
        data_query[PERIODO_FIELD] = periodo.strip().upper()
    if datas_list:
        data_query[IMPORT_DATE_FIELD] = {"$in": datas_list}

    motorista_norm = (motorista or "").strip() or "(sem motorista)"
    base_norm = (base or "").strip() or "(sem base)"
    
    # Buscar pedidos da entrada no galpão que devem ser excluídos
    pedidos_excluir = {}
    try:
        col_entrada = db[COLLECTION_ENTRADA_GALPAO]
        header_entrada_doc = col_entrada.find_one(
            {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
            sort=[("_id", 1)],
        )
        if header_entrada_doc:
            header_entrada = list(header_entrada_doc.get("values", []))
            header_entrada_id = header_entrada_doc["_id"]
            header_entrada_norm = [_normalize_text(h) for h in header_entrada]
            col_jms_entrada_norm = _normalize_text(COL_JMS_ENTRADA)
            col_tipo_bipagem_entrada_norm = _normalize_text(COL_TIPO_BIPAGEM)
            col_tempo_digitalizacao_norm = _normalize_text(COL_TEMPO_DIGITALIZACAO)
            idx_jms_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_jms_entrada_norm == h), -1)
            idx_tipo_bipagem_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tipo_bipagem_entrada_norm == h), -1)
            idx_tempo_digitalizacao_entrada = next((i for i, h in enumerate(header_entrada_norm) if col_tempo_digitalizacao_norm == h), -1)
            
            if idx_jms_entrada >= 0 and idx_tipo_bipagem_entrada >= 0:
                entrada_query = {**q_user, "_id": {"$ne": header_entrada_id}}
                if datas_list:
                    entrada_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
                # Ordenar por _id desc para pegar o mais recente primeiro
                cursor_entrada = col_entrada.find(entrada_query, {"values": 1}).sort("_id", -1)
                # Agrupar por JMS para pegar apenas o mais recente
                candidatos_por_jms = {}
                for doc_entrada in cursor_entrada:
                    vals_entrada = doc_entrada.get("values") or []
                    jms_entrada = _get(vals_entrada, idx_jms_entrada) if idx_jms_entrada >= 0 else ""
                    tipo_bipagem = _get(vals_entrada, idx_tipo_bipagem_entrada) if idx_tipo_bipagem_entrada >= 0 else ""
                    if jms_entrada and tipo_bipagem:
                        jms_entrada_norm = str(jms_entrada).strip()
                        tipo_bipagem_norm = _normalize_text(str(tipo_bipagem))
                        if tipo_bipagem_norm == _normalize_text(TIPO_BIPAGEM_EXCLUIR):
                            # Agrupar por JMS (já ordenado por _id desc, então o primeiro será o mais recente)
                            if jms_entrada_norm not in candidatos_por_jms:
                                tempo_digitalizacao = _get(vals_entrada, idx_tempo_digitalizacao_entrada) if idx_tempo_digitalizacao_entrada >= 0 else None
                                pedidos_excluir[jms_entrada_norm] = tempo_digitalizacao
                                candidatos_por_jms[jms_entrada_norm] = True
    except Exception:
        pass

    docs = []
    cursor = col.find(data_query, {"values": 1, IMPORT_DATE_FIELD: 1, "createdAt": 1})
    for doc in cursor:
        vals = doc.get("values") or []
        if _normalize_text(_get(vals, idx_motorista)) != _normalize_text(motorista_norm):
            continue
        if _normalize_text(_get(vals, idx_base)) != _normalize_text(base_norm):
            continue
        marca_norm = _normalize_marca_val(_get(vals, idx_marca))
        if marca_norm not in MARCAS_ENTREGUE:
            continue
        if cidades_list and idx_cidade >= 0:
            cidade_norm = _normalize_text((_get(vals, idx_cidade) or "").strip()) if (_get(vals, idx_cidade) or "").strip() else ""
            if cidade_norm not in cidades_list:
                continue
        # Excluir pedidos que estão na entrada no galpão
        # Mas só excluir se o horário de saída não existir ou for antes/igual ao tempo de digitalização
        if idx_jms >= 0:
            jms_value = str(_get(vals, idx_jms) or "").strip()
            if jms_value and jms_value in pedidos_excluir:
                tempo_digitalizacao = pedidos_excluir[jms_value]
                horario_saida = _get(vals, idx_horario_saida) if idx_horario_saida >= 0 else None
                
                # Se não há horário de saída, excluir (pedido não saiu para entrega)
                # Se há horário de saída, comparar com tempo de digitalização
                if horario_saida is None or not str(horario_saida).strip():
                    continue
                else:
                    # Comparar horários: se horário de saída é depois do tempo de digitalização, NÃO excluir
                    comparacao = _compare_times(horario_saida, tempo_digitalizacao)
                    if comparacao is None:
                        # Se não conseguiu comparar, excluir por segurança
                        continue
                    elif comparacao <= 0:
                        # Horário de saída é antes ou igual ao tempo de digitalização, excluir
                        continue
                    # Se comparacao > 0 (horário de saída é depois), não excluir e continuar processando
        c = doc.get("createdAt")
        docs.append({
            "_id": str(doc["_id"]),
            "values": vals,
            "createdAt": c.isoformat() if c else None,
            "importDate": doc.get(IMPORT_DATE_FIELD),
        })

    return {"data": docs, "header": header}


@router.get("/datas")
def listar_datas_importacao(
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """Retorna as datas de importação existentes na coleção SLA, ordenadas da mais recente."""
    db = get_db()
    col = db[COLLECTION]
    q = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
    datas = col.distinct(IMPORT_DATE_FIELD, q)
    datas = sorted([d for d in datas if d], reverse=True)
    return {"datas": datas}


@router.get("")
def listar_sla(
    page: int = 1,
    per_page: int = 100,
    datas: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna registros SLA com paginação (suporta grandes volumes).
    datas: opcional, vírgulas (ex: 2026-02-08,2026-02-09).
    page=1 retorna também o header.
    """
    per_page = min(max(1, per_page), 500)
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}

    query = dict(q_user)
    datas_list = _parse_csv_param(datas)
    if datas_list:
        query[IMPORT_DATE_FIELD] = {"$in": datas_list}

    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    header = list(header_doc.get("values", [])) if header_doc else []
    header_id = header_doc["_id"] if header_doc else None

    data_query = dict(q_user)
    if header_id:
        data_query["_id"] = {"$ne": header_id}
    if datas_list:
        data_query[IMPORT_DATE_FIELD] = {"$in": datas_list}
    total = col.count_documents(data_query)

    if total == 0:
        return {"data": [], "total": 0, "header": header if page == 1 else None}

    skip = (page - 1) * per_page
    cursor = col.find(data_query).sort("_id", 1).skip(skip).limit(per_page)

    docs = []
    for doc in cursor:
        c = doc.get("createdAt")
        # Garantir serialização correta de datetime
        if c:
            if hasattr(c, "isoformat"):
                created_at = c.isoformat()
            elif isinstance(c, str):
                created_at = c
            else:
                created_at = str(c)
        else:
            created_at = None
        
        docs.append({
            "_id": str(doc["_id"]),
            "values": doc.get("values", []),
            "createdAt": created_at,
            "importDate": doc.get(IMPORT_DATE_FIELD),
        })

    result = {"data": docs, "total": total, "header": header if page == 1 else None}
    # Debug: verificar se dados estão sendo retornados (usar sys.stderr para aparecer no executável)
    import sys
    sys.stderr.write(f"[DEBUG] listar_sla - user_id: {user_id}, total: {total}, docs_count: {len(docs)}, page: {page}\n")
    sys.stderr.flush()
    return result


@router.delete("")
def deletar_todos(
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """Remove todos os documentos da coleção SLA do usuário."""
    db = get_db()
    col = db[COLLECTION]
    result = col.delete_many({USER_ID_FIELD: user_id})
    return {"deleted": result.deleted_count}


@router.delete("/{doc_id}")
def deletar_linha(
    doc_id: str,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """Remove um documento pelo _id."""
    try:
        oid = ObjectId(doc_id)
    except InvalidId:
        raise HTTPException(status_code=400, detail="ID inválido.")
    db = get_db()
    col = db[COLLECTION]
    result = col.delete_one({"_id": oid, USER_ID_FIELD: user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro não encontrado.")
    return {"deleted": 1}
