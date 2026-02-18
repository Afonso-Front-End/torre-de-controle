"""
Rotas: lista de telefones – recebe arquivo Excel, processa, salva na coleção, retorna dados.
Delete exige senha do usuário logado e grava histórico (não exibido).
"""
from datetime import datetime, timezone
from io import BytesIO

from bson.errors import InvalidId
from bson.objectid import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from database import get_db, USER_ID_FIELD
from limiter import limiter
from routers.auth import require_user_id
from security import verify_password
from table_ids import require_table_id

router = APIRouter(prefix="/lista-telefones", tags=["lista-telefones"])
COLLECTION = "lista_telefones"
HISTORY_COLLECTION = "lista_telefones_delete_history"
USUARIOS_COLLECTION = "usuarios"
CHUNK_SIZE = 1000  # inserir no MongoDB em lotes (insert_many)
IMPORT_DATE_FIELD = "importDate"
HEADER_FLAG = "isHeader"


def _parse_datas_query(datas: str | None) -> list[str] | None:
    if not datas or not str(datas).strip():
        return None
    return [d.strip() for d in str(datas).split(",") if d.strip()]


def _indices_contato(header_values: list) -> tuple[int | None, int | None, int | None]:
    """
    Obtém os índices das colunas Motorista, HUB e Contato no cabeçalho.
    Retorna (motorista_idx, hub_idx, contato_idx); None se não encontrar.
    """
    if not header_values:
        return None, None, None
    norm = lambda s: (s or "").strip().upper()
    motorista_idx = hub_idx = contato_idx = None
    for i, v in enumerate(header_values):
        n = norm(str(v))
        if n == "MOTORISTA":
            motorista_idx = i
        elif n == "HUB":
            hub_idx = i
        elif n == "CONTATO":
            contato_idx = i
    return motorista_idx, hub_idx, contato_idx


class UpdateHubBody(BaseModel):
    col_index: int
    valor_atuais: list[str]  # valores brutos a substituir (ex.: ["   BNU -SC", "BNU -SC"])
    valor_novo: str


class DeleteListaBody(BaseModel):
    """Senha do usuário logado para confirmar delete."""
    senha: str = Field(..., min_length=1)


class UpdateContatoBody(BaseModel):
    """Atualiza o campo Contato de um registro por _id."""
    doc_id: str = Field(..., min_length=1)
    contato: str = Field(default="")


def excel_para_linhas(contents: bytes) -> list:
    """Lê a primeira planilha do Excel e retorna lista de linhas (cada linha = lista de strings)."""
    wb = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        rows.append(cells)
    wb.close()
    return rows


@router.post("")
@limiter.limit("20/minute")
async def salvar_lista(request: Request, file: UploadFile = File(...), user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """
    Recebe um arquivo Excel (.xlsx). Modo incremental: não apaga dados anteriores.
    Cada linha é gravada com importDate (data do envio) para filtro por data.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo não informado.")
    ext = (file.filename or "").lower()
    if not ext.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx")

    contents = await file.read()
    try:
        rows = excel_para_linhas(contents)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o Excel: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="O arquivo está vazio ou não tem dados na primeira planilha.")

    db = get_db()
    col = db[COLLECTION]
    now = datetime.now(timezone.utc)
    import_date_str = now.strftime("%Y-%m-%d")
    q_user = {USER_ID_FIELD: user_id}

    # Primeira vez para este usuário: gravar primeira linha como cabeçalho (sem importDate).
    if col.count_documents(q_user) == 0:
        col.insert_one({**q_user, "values": list(rows[0]), HEADER_FLAG: True})
        rows = rows[1:] if len(rows) > 1 else []

    docs = []
    for i in range(0, len(rows), CHUNK_SIZE):
        chunk = rows[i : i + CHUNK_SIZE]
        batch = [{**q_user, "values": row, "createdAt": now, IMPORT_DATE_FIELD: import_date_str} for row in chunk]
        result = col.insert_many(batch)
        for row, oid in zip(chunk, result.inserted_ids):
            docs.append({
                "_id": str(oid),
                "values": row,
                "createdAt": now.isoformat(),
                "importDate": import_date_str,
            })
    
    # Debug: verificar se dados foram salvos corretamente
    import sys
    sys.stderr.write(f"[DEBUG] salvar_lista_telefones - user_id: {user_id}, saved: {len(docs)}, total_no_banco: {col.count_documents({USER_ID_FIELD: user_id})}\n")
    sys.stderr.flush()
    
    return {"saved": len(docs), "data": docs}


@router.get("/datas")
def listar_datas_importacao(user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """Retorna as datas de importação (importDate) existentes na coleção, ordenadas da mais recente."""
    db = get_db()
    col = db[COLLECTION]
    q = {USER_ID_FIELD: user_id, IMPORT_DATE_FIELD: {"$exists": True}}
    datas = col.distinct(IMPORT_DATE_FIELD, q)
    datas = sorted([d for d in datas if d], reverse=True)
    return {"datas": datas}


@router.get("")
def listar_telefones(
    datas: str | None = None,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Retorna os documentos da coleção lista_telefones.
    Primeiro elemento = cabeçalho (se existir); restante = linhas de dados.
    datas: opcional, vírgulas (ex: 2026-02-08,2026-02-09) – exibe apenas registros dessas datas de envio.
    """
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}
    datas_list = _parse_datas_query(datas)
    data_query = dict(q_user)
    if datas_list:
        data_query[IMPORT_DATE_FIELD] = {"$in": datas_list}

    # Cabeçalho: doc com isHeader ou (retrocompat) primeiro doc sem importDate, do usuário.
    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    header_id = header_doc["_id"] if header_doc else None
    docs = []
    if header_doc:
        c = header_doc.get("createdAt")
        # Garantir serialização correta de datetime
        if c:
            if hasattr(c, "isoformat"):
                created_at = c.isoformat()
            elif isinstance(c, str):
                created_at = c
            else:
                created_at = str(c)
        else:
            created_at = None
        
        docs.append({
            "_id": str(header_doc["_id"]),
            "values": header_doc.get("values", []),
            "createdAt": created_at,
            "importDate": None,
        })
    # Dados: do usuário; se filtro por datas = apenas docs com importDate em datas_list; senão = todos exceto o header.
    if datas_list:
        query_data = {**q_user, IMPORT_DATE_FIELD: {"$in": datas_list}}
    else:
        query_data = {**q_user} if not header_id else {**q_user, "_id": {"$ne": header_id}}
    for doc in col.find(query_data).sort("_id", 1):
        c = doc.get("createdAt")
        # Garantir serialização correta de datetime
        if c:
            if hasattr(c, "isoformat"):
                created_at = c.isoformat()
            elif isinstance(c, str):
                created_at = c
            else:
                created_at = str(c)
        else:
            created_at = None
        
        docs.append({
            "_id": str(doc["_id"]),
            "values": doc.get("values", []),
            "createdAt": created_at,
            "importDate": doc.get(IMPORT_DATE_FIELD),
        })
    result = {"data": docs}
    # Debug: verificar se dados estão sendo retornados (usar sys.stderr para aparecer no executável)
    import sys
    sys.stderr.write(f"[DEBUG] listar_telefones - user_id: {user_id}, docs_count: {len(docs)}, datas: {datas}\n")
    sys.stderr.flush()
    return result


@router.get("/contato")
def buscar_contato_motorista_base(
    motorista: str = "",
    base: str = "",
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Busca o contato (telefone) na lista_telefones onde Motorista = motorista e HUB = base.
    Responsável pela entrega = Motorista, Base de entrega = HUB.
    Retorna { contato, _id } para preencher o modal; _id usado para atualizar depois.
    """
    db = get_db()
    col = db[COLLECTION]
    q_user = {USER_ID_FIELD: user_id}
    header_doc = col.find_one(
        {**q_user, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_doc:
        return {"contato": "", "_id": None}
    header_values = header_doc.get("values") or []
    motorista_idx, hub_idx, contato_idx = _indices_contato(header_values)
    if motorista_idx is None or hub_idx is None or contato_idx is None:
        return {"contato": "", "_id": None}
    header_id = header_doc["_id"]
    query = {
        USER_ID_FIELD: user_id,
        "_id": {"$ne": header_id},
        f"values.{motorista_idx}": motorista.strip(),
        f"values.{hub_idx}": base.strip(),
    }
    doc = col.find_one(query, sort=[("_id", 1)])
    if not doc:
        return {"contato": "", "_id": None}
    values = doc.get("values") or []
    contato = (values[contato_idx] if contato_idx < len(values) else "") or ""
    return {"contato": contato.strip(), "_id": str(doc["_id"])}


@router.patch("/contato")
def atualizar_contato_motorista_base(
    body: UpdateContatoBody,
    user_id: str = Depends(require_user_id),
    table_id: int = Depends(require_table_id),
):
    """
    Atualiza o campo Contato (values[contato_idx]) do documento com _id = doc_id.
    """
    try:
        oid = ObjectId(body.doc_id)
    except InvalidId:
        raise HTTPException(status_code=400, detail="ID do registro inválido.")
    db = get_db()
    col = db[COLLECTION]
    header_doc = col.find_one(
        {USER_ID_FIELD: user_id, "$or": [{HEADER_FLAG: True}, {IMPORT_DATE_FIELD: {"$exists": False}}]},
        sort=[("_id", 1)],
    )
    if not header_doc:
        raise HTTPException(status_code=404, detail="Cabeçalho da lista de telefones não encontrado.")
    header_values = header_doc.get("values") or []
    _, _, contato_idx = _indices_contato(header_values)
    if contato_idx is None:
        raise HTTPException(status_code=400, detail="Coluna Contato não encontrada no cabeçalho.")
    result = col.update_one(
        {"_id": oid, USER_ID_FIELD: user_id},
        {"$set": {f"values.{contato_idx}": body.contato.strip()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Registro não encontrado.")
    return {"updated": 1}


def _verificar_senha_e_obter_usuario(db, user_id: str, senha: str) -> dict:
    """Carrega usuário por ID e verifica senha. Levanta 401 se inválido."""
    user = db[USUARIOS_COLLECTION].find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=401, detail="Usuário não encontrado.")
    if not verify_password(senha, user["senha_hash"]):
        raise HTTPException(status_code=401, detail="Senha incorreta.")
    return user


def _registrar_delete_history(db, user_id: str, user_name: str, action: str, deleted_count: int | None = None, row_id: str | None = None):
    """Grava registro no histórico de delete (não exibido na UI)."""
    now = datetime.now(timezone.utc)
    doc = {
        "userId": user_id,
        "userName": user_name,
        "action": action,
        "createdAt": now,
    }
    if deleted_count is not None:
        doc["deletedCount"] = deleted_count
    if row_id is not None:
        doc["rowId"] = row_id
    db[HISTORY_COLLECTION].insert_one(doc)


@router.delete("")
@limiter.limit("20/minute")
def deletar_lista(request: Request, body: DeleteListaBody, user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """
    Remove todos os documentos da coleção lista_telefones.
    Exige Authorization: Bearer <token> e body com senha do usuário logado.
    Registra no histórico de delete (não exibido).
    """
    db = get_db()
    user = _verificar_senha_e_obter_usuario(db, user_id, body.senha)
    col = db[COLLECTION]
    result = col.delete_many({USER_ID_FIELD: user_id})
    _registrar_delete_history(
        db, user_id, user.get("nome", ""), "delete_all", deleted_count=result.deleted_count
    )
    return {"deleted": result.deleted_count}


@router.delete("/{doc_id}")
@limiter.limit("30/minute")
def deletar_linha(request: Request, doc_id: str, body: DeleteListaBody, user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """
    Remove um documento da coleção lista_telefones pelo _id.
    Exige Authorization: Bearer <token> e body com senha do usuário logado.
    Registra no histórico de delete (não exibido).
    """
    try:
        oid = ObjectId(doc_id)
    except InvalidId:
        raise HTTPException(status_code=400, detail="ID inválido.")
    db = get_db()
    user = _verificar_senha_e_obter_usuario(db, user_id, body.senha)
    col = db[COLLECTION]
    result = col.delete_one({"_id": oid, USER_ID_FIELD: user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro não encontrado.")
    _registrar_delete_history(db, user_id, user.get("nome", ""), "delete_row", row_id=doc_id)
    return {"deleted": 1}


@router.patch("")
def atualizar_coluna_hub(body: UpdateHubBody, user_id: str = Depends(require_user_id), table_id: int = Depends(require_table_id)):
    """
    Atualiza todos os registros em que a coluna (col_index) tem um dos valor_atuais,
    substituindo por valor_novo (ex.: ["   BNU -SC", "BNU -SC"] → "BNU SC").
    """
    db = get_db()
    col = db[COLLECTION]
    if body.col_index < 0:
        raise HTTPException(status_code=400, detail="col_index inválido.")
    if not body.valor_atuais:
        return {"updated": 0}
    key = f"values.{body.col_index}"
    total = 0
    for valor in body.valor_atuais:
        result = col.update_many(
            {USER_ID_FIELD: user_id, key: valor},
            {"$set": {key: body.valor_novo}},
        )
        total += result.modified_count
    return {"updated": total}
