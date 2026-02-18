"""
Lógica de negócio: resultados da consulta.
Constantes, helpers Excel e processamento para coleções base/motorista.
"""
from datetime import datetime, timezone
from io import BytesIO

from bson.objectid import ObjectId
from openpyxl import load_workbook
from pymongo.errors import PyMongoError

# Coleções MongoDB
COLLECTION_PEDIDOS = "pedidos"
COLLECTION_PEDIDOS_STATUS = "pedidos_com_status"
COLLECTION_BASE = "base"
COLLECTION_MOTORISTA = "motorista"
COLLECTION_USUARIOS = "usuarios"

# Colunas que vêm da linha do pedido (por nome de coluna no header de pedidos)
COLUNAS_PEDIDO = [
    "Base de entrega",
    "CEP destino",
    "Complemento",
    "Destinatário",
    "Cidade Destino",
    "3 Segmentos",
    "Distrito destinatário",
    "Marca de assinatura",
    "Horário da entrega",
    "PDD de Entrega",
]

# Prefixos do "Correio de coleta ou entrega" que indicam motorista (TAC, MEI, ETC)
MOTORISTA_PREFIXOS_CORREIO_DEFAULT = ["TAC", "MEI", "ETC"]

# Ordem exata dos campos ao gravar na coleção motorista (sem createdAt)
ORDEM_CAMPOS_MOTORISTA = [
    "Número de pedido JMS",
    "Correio de coleta ou entrega",
    "Base de entrega",
    "Tipo de bipagem",
    "Tempo de digitalização",
    "Marca de assinatura",
    "Dias sem movimentação",
    "CEP destino",
    "Complemento",
    "Destinatário",
    "Cidade Destino",
    "Distrito destinatário",
    "PDD de Entrega",
    "Status",
]

# Valores de "Marca de assinatura" que disparam atualização do doc na coleção motorista
MARCA_ENTREGUE = ["Recebimento com assinatura normal", "Assinatura de devolução"]
# Valor que indica não entrega (usado para filtrar e para recalcular dias parados)
MARCA_NAO_ENTREGUE = "Não entregue"

JMS_FIELD = "Número de pedido JMS"
MARCA_FIELD = "Marca de assinatura"
TEMPO_DIG_FIELD = "Tempo de digitalização"

# Data do envio (YYYY-MM-DD) para filtrar por data de importação na coleção motorista/base
IMPORT_DATE_FIELD = "importDate"


def sanitize_cell(value):
    """Converte célula Excel para string consistente."""
    if value is None:
        return ""
    if isinstance(value, (datetime,)):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""
    return str(value).strip()


def excel_para_linhas(contents: bytes) -> list:
    """Lê a primeira planilha e retorna lista de linhas (cada linha = lista de strings sanitizadas)."""
    wb = load_workbook(filename=BytesIO(contents), read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = []
    max_row = ws.max_row if ws.max_row else 0
    max_col = ws.max_column if ws.max_column else 1
    for row_idx in range(1, max_row + 1):
        row = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row.append(sanitize_cell(cell.value))
        rows.append(row)
    wb.close()
    return rows


def _normalize_header(h):
    """Normaliza nome de coluna para comparação."""
    if h is None:
        return ""
    return str(h).strip().lower()


def idx_coluna(header, nome):
    """Retorna índice da coluna pelo nome (normalizado)."""
    for i, h in enumerate(header):
        if _normalize_header(h) == _normalize_header(nome):
            return i
    return -1


def idx_digitalizador(header: list) -> int:
    """Retorna índice da coluna Digitalizador (nome exato ou que contenha 'digitalizador')."""
    idx = idx_coluna(header, "Digitalizador")
    if idx >= 0:
        return idx
    for i, h in enumerate(header):
        if "digitalizador" in _normalize_header(h):
            return i
    return -1


def idx_numero_pedido_jms(header):
    for i, h in enumerate(header):
        t = _normalize_header(h)
        if "número de pedido jms" in t or "numero de pedido jms" in t:
            return i
    return -1


def idx_tempo_digitalizacao(header):
    n = "tempo de digitalização"
    n2 = "tempo de digitalizacao"
    for i, h in enumerate(header):
        t = _normalize_header(h)
        if n in t or n2 in t:
            return i
    return -1


def _parse_tempo(v):
    """Converte valor para datetime. Retorna None se inválido."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=timezone.utc) if not v.tzinfo else v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:26], fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _dias_desde(tempo_dt):
    """Retorna número de dias desde tempo_dt até hoje (UTC)."""
    if tempo_dt is None:
        return None
    agora = datetime.now(timezone.utc)
    if tempo_dt.tzinfo is None:
        tempo_dt = tempo_dt.replace(tzinfo=timezone.utc)
    delta = agora - tempo_dt
    return max(0, delta.days)


def dias_sem_movimentacao_motorista(doc: dict) -> int | None:
    """
    Recalcula 'Dias sem movimentação' ao listar motorista.
    Sempre usa Tempo de digitalização (última bipagem): dias desde essa data até hoje.
    """
    tempo_str = doc.get(TEMPO_DIG_FIELD)
    tempo_dt = _parse_tempo(tempo_str)
    return _dias_desde(tempo_dt)


def motorista_prefixos_list(config: dict) -> list:
    """Lista de prefixos do Correio que indicam motorista (config.motorista_prefixos_correio)."""
    raw = config.get("motorista_prefixos_correio")
    if not raw:
        return list(MOTORISTA_PREFIXOS_CORREIO_DEFAULT)
    if isinstance(raw, str):
        raw = [s.strip() for s in raw.split(",") if s.strip()]
    if not isinstance(raw, list):
        return list(MOTORISTA_PREFIXOS_CORREIO_DEFAULT)
    return [str(p).strip() for p in raw if p is not None and str(p).strip()]


def salvar_prefixos_motorista_no_utilizador(db, user_id: str, prefixos: list) -> None:
    """Persiste a lista de prefixos na config do utilizador."""
    if not prefixos:
        return
    lista = [str(p).strip() for p in prefixos if p is not None and str(p).strip()]
    try:
        db[COLLECTION_USUARIOS].update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"config.motorista_prefixos_correio": lista}},
        )
    except PyMongoError:
        pass


def exigir_digitalizador_motorista(config: dict) -> bool:
    """Se True, exige Digitalizador preenchido para considerar como motorista."""
    return config.get("exigir_digitalizador_motorista", True)


def _valor_tem_prefixo(valor: str, prefixos: list) -> bool:
    """True se valor (trimmed) começa com um dos prefixos (case-insensitive)."""
    if not valor or not prefixos:
        return False
    s = str(valor).strip().upper()
    return any(s.startswith(str(p).strip().upper()) for p in prefixos if p)


def eh_motorista(digitalizador: str, correio: str, prefixos: list, exigir_digitalizador: bool) -> bool:
    """
    True se deve enviar para coleção motorista.
    - Ambos preenchidos → priorizar Correio (prefixo).
    - Apenas Digitalizador → filtrar pelos prefixos no Digitalizador.
    - Apenas Correio → prefixo no Correio (e, se exigir_digitalizador, rejeita).
    - Nenhum preenchido → não é motorista.
    """
    d = str(digitalizador or "").strip()
    c = str(correio or "").strip()
    if not prefixos:
        return False
    if d and c:
        return _valor_tem_prefixo(c, prefixos)
    if d and not c:
        return _valor_tem_prefixo(d, prefixos)
    if c and not d:
        if exigir_digitalizador:
            return False
        return _valor_tem_prefixo(c, prefixos)
    return False


def row_to_motorista_update(header: list, row: list) -> dict:
    """Constrói dicionário de atualização para motorista a partir do cabeçalho e da linha."""
    update = {}
    for field_name in ORDEM_CAMPOS_MOTORISTA:
        idx = idx_coluna(header, field_name)
        if idx >= 0 and idx < len(row):
            val = row[idx]
            update[field_name] = val
    return update


def processar_e_gravar(
    db,
    numeros_jms: list,
    colecao: str,
    user_id: str,
    motorista_prefixos: list = None,
    motorista_exigir_digitalizador: bool = True,
) -> dict:
    """
    Lógica central: para cada numero_jms, busca em pedidos_com_status e opcionalmente
    em pedidos; monta o documento e grava na coleção com userId.
    Para coleção motorista, só grava quando atende ao critério (prefixos, etc.).
    Retorna {"saved", "skipped", "rejected_tipo_bipagem"}.
    """
    from database import USER_ID_FIELD

    if motorista_prefixos is None and colecao == "motorista":
        motorista_prefixos = list(MOTORISTA_PREFIXOS_CORREIO_DEFAULT)
    col_pedidos = db[COLLECTION_PEDIDOS]
    col_status = db[COLLECTION_PEDIDOS_STATUS]
    col_dest_name = COLLECTION_BASE if colecao == "base" else COLLECTION_MOTORISTA
    col_dest = db[col_dest_name]
    q_user = {USER_ID_FIELD: user_id}

    first_pedidos = col_pedidos.find_one({**q_user, "isHeader": True}) or col_pedidos.find_one(q_user, sort=[("_id", 1)])
    header_pedidos = list(first_pedidos.get("values", [])) if first_pedidos else []
    idx_jms_pedidos = idx_numero_pedido_jms(header_pedidos) if header_pedidos else -1
    id_header_pedidos = first_pedidos["_id"] if first_pedidos else None
    map_col_pedido = {}
    if first_pedidos:
        for nome in COLUNAS_PEDIDO:
            i = idx_coluna(header_pedidos, nome)
            if i >= 0:
                map_col_pedido[nome] = i

    first_status = col_status.find_one({**q_user, "isHeader": True}) or col_status.find_one(q_user, sort=[("_id", 1)])
    header_status = list(first_status.get("values", [])) if first_status else []
    id_header_status = first_status["_id"] if first_status else None
    idx_jms_status = idx_numero_pedido_jms(header_status)
    if idx_jms_status < 0:
        return {"saved": 0, "skipped": 0, "rejected_tipo_bipagem": 0}
    idx_tempo_status = idx_tempo_digitalizacao(header_status)
    idx_tipo_bipagem = -1
    for i, h in enumerate(header_status):
        t = _normalize_header(h)
        if "tipo de bipagem" in t or "tipo bipagem" in t:
            idx_tipo_bipagem = i
            break
    idx_correio = idx_coluna(header_status, "Correio de coleta ou entrega")
    idx_dig = idx_digitalizador(header_status)

    saved = 0
    skipped = 0
    rejected_tipo_bipagem = 0
    # Data local do servidor (não UTC), para coincidir com «hoje» no frontend e evitar dia diferente na BD
    import_date_str = datetime.now().date().strftime("%Y-%m-%d")

    for numero_jms in numeros_jms:
        if not numero_jms:
            continue
        if col_dest.find_one({**q_user, "Número de pedido JMS": numero_jms}):
            skipped += 1
            continue
        doc_status = None
        if id_header_status is not None:
            doc_status = col_status.find_one(
                {**q_user, "isHeader": {"$ne": True}, "_id": {"$ne": id_header_status}, f"values.{idx_jms_status}": numero_jms},
                sort=[("_id", 1)],
            )
        if not doc_status:
            continue
        values_status = doc_status.get("values") or []
        doc_pedido = None
        if first_pedidos and id_header_pedidos and idx_jms_pedidos >= 0:
            doc_pedido = col_pedidos.find_one(
                {**q_user, "isHeader": {"$ne": True}, "_id": {"$ne": id_header_pedidos}, f"values.{idx_jms_pedidos}": numero_jms},
                sort=[("_id", 1)],
            )
        values_pedido = (doc_pedido or {}).get("values") or []
        row = {}
        for nome, idx in map_col_pedido.items():
            row[nome] = values_pedido[idx] if idx < len(values_pedido) else ""
        tempo_dt = None
        if idx_tempo_status >= 0 and idx_tempo_status < len(values_status):
            tempo_dt = _parse_tempo(values_status[idx_tempo_status])
        tipo_bipagem = ""
        if idx_tipo_bipagem >= 0 and idx_tipo_bipagem < len(values_status):
            tipo_bipagem = values_status[idx_tipo_bipagem] or ""
        correio_coleta_entrega = ""
        if idx_correio >= 0 and idx_correio < len(values_status):
            correio_coleta_entrega = values_status[idx_correio] or ""
        digitalizador = ""
        if idx_dig >= 0 and idx_dig < len(values_status):
            digitalizador = values_status[idx_dig] or ""
        row["Número de pedido JMS"] = numero_jms
        row["Tipo de bipagem"] = tipo_bipagem
        row["Tempo de digitalização"] = values_status[idx_tempo_status] if idx_tempo_status >= 0 and idx_tempo_status < len(values_status) else ""
        row["Correio de coleta ou entrega"] = correio_coleta_entrega
        row["Status"] = ""
        row["Dias sem movimentação"] = _dias_desde(tempo_dt)
        if colecao == "motorista":
            if not eh_motorista(digitalizador, correio_coleta_entrega, motorista_prefixos, motorista_exigir_digitalizador):
                rejected_tipo_bipagem += 1
                continue
            if not (row.get("Correio de coleta ou entrega") or "").strip() and (digitalizador or "").strip():
                row["Correio de coleta ou entrega"] = (digitalizador or "").strip()
        doc_dest = {k: row.get(k, "") for k in ORDEM_CAMPOS_MOTORISTA}
        doc_dest[USER_ID_FIELD] = user_id
        doc_dest[IMPORT_DATE_FIELD] = import_date_str
        try:
            col_dest.insert_one(doc_dest)
            saved += 1
        except PyMongoError:
            pass

    return {"saved": saved, "skipped": skipped, "rejected_tipo_bipagem": rejected_tipo_bipagem}
